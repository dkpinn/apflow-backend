from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

import app.services.bank_statement_extraction as extraction
import app.services.bank_statement_service as facade
from app.services.extraction_foundation import detect_source_format
from app.services.extractor_registry import select_bank_cash_extractor


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_xlsx_debit_credit_rows_support_native_dates_numbers_and_hashes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(["Account statement"])
    sheet.append([])
    sheet.append(["Date", "Value Date", "Description", "Reference", "Debit", "Credit", "Balance"])
    sheet.append([date(2026, 1, 1), datetime(2026, 1, 2, 9, 30), "Supplier payment", "INV-100", 100, None, 900])
    sheet.append([date(2026, 1, 3), None, "Customer receipt", "RCPT-1", None, 250.5, 1150.5])

    header, lines = extraction.parse_xlsx_statement(
        _workbook_bytes(workbook),
        bank_account_id="bank-1",
        currency="ZAR",
    )

    assert header["source_format"] == "xlsx"
    assert header["parser_strategy"] == "deterministic_xlsx"
    assert header["opening_balance"] == 1000.0
    assert header["closing_balance"] == 1150.5
    assert header["raw_extraction"]["sheet_name"] == "Transactions"
    assert header["raw_extraction"]["header_row"] == 3
    assert lines[0].line_date == "2026-01-01"
    assert lines[0].value_date == "2026-01-02"
    assert lines[0].signed_amount == facade.money("-100")
    assert lines[1].signed_amount == facade.money("250.50")
    assert lines[0].transaction_hash


def test_xlsx_signed_amount_reuses_csv_normalization_and_fingerprint():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Description", "Reference", "Amount", "Balance"])
    sheet.append([date(2026, 2, 1), "Monthly fee", "FEE-1", -50, 950])
    xlsx_bytes = _workbook_bytes(workbook)
    csv_bytes = b"Date,Description,Reference,Amount,Balance\n2026-02-01,Monthly fee,FEE-1,-50,950\n"

    _xlsx_header, xlsx_lines = extraction.parse_xlsx_statement(xlsx_bytes, bank_account_id="bank-1")
    _csv_header, csv_lines = extraction.parse_csv_statement(csv_bytes, bank_account_id="bank-1")

    assert xlsx_lines[0].debit_amount == facade.money("50")
    assert xlsx_lines[0].credit_amount == facade.money("0")
    assert xlsx_lines[0].transaction_hash == csv_lines[0].transaction_hash


def test_xlsx_selects_strongest_visible_sheet_and_uses_workbook_order_for_ties():
    workbook = Workbook()
    hidden = workbook.active
    hidden.title = "Hidden"
    hidden.sheet_state = "hidden"
    hidden.append(["Date", "Description", "Reference", "Debit", "Credit", "Balance", "Currency"])
    hidden.append([date(2026, 1, 1), "Hidden row", "H-1", None, 1, 1, "ZAR"])

    first = workbook.create_sheet("First")
    first.append(["Date", "Description", "Amount"])
    first.append([date(2026, 1, 1), "Weak row", 10])

    second = workbook.create_sheet("Second")
    second.append(["Statement title"])
    second.append(["Date", "Description", "Reference", "Debit", "Credit", "Balance"])
    second.append([date(2026, 1, 2), "Selected row", "S-1", None, 20, 1020])

    tied = workbook.create_sheet("Tied")
    tied.append(["Date", "Description", "Reference", "Debit", "Credit", "Balance"])
    tied.append([date(2026, 1, 3), "Later tied row", "T-1", None, 30, 1050])

    header, lines = extraction.parse_xlsx_statement(
        _workbook_bytes(workbook),
        bank_account_id="bank-1",
    )

    assert header["raw_extraction"]["sheet_name"] == "Second"
    assert [line.description for line in lines] == ["Selected row"]


def test_xlsx_skips_empty_rows_and_rejects_invalid_workbooks():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Description", "Amount"])
    sheet.append([])
    sheet.append([date(2026, 3, 1), "Deposit", 125])

    _header, lines = extraction.parse_xlsx_statement(
        _workbook_bytes(workbook),
        bank_account_id="bank-1",
    )
    assert len(lines) == 1

    with pytest.raises(ValueError, match="Could not read XLSX"):
        extraction.parse_xlsx_statement(b"not an xlsx workbook", bank_account_id="bank-1")

    no_header = Workbook()
    no_header.active.append(["Monthly bank statement"])
    no_header.active.append(["Nothing", "Recognizable"])
    with pytest.raises(ValueError, match="recognizable transaction header"):
        extraction.parse_xlsx_statement(_workbook_bytes(no_header), bank_account_id="bank-1")


def test_xlsx_detection_routing_and_legacy_xls_rejection():
    assert detect_source_format("statement.xlsx", None) == "xlsx"
    assert detect_source_format("statement", XLSX_MIME) == "xlsx"
    assert detect_source_format("statement.xls", "application/vnd.ms-excel") == "xls"
    assert detect_source_format("statement.csv", "application/vnd.ms-excel") == "csv"

    selection = select_bank_cash_extractor(
        account_type="bank",
        filename="statement.xlsx",
        mime_type=XLSX_MIME,
    )
    assert selection.source_format == "xlsx"
    assert selection.parser_strategy == "deterministic_xlsx"

    with pytest.raises(ValueError, match=r"export the statement as \.xlsx or \.csv"):
        extraction.extract_statement(
            b"legacy workbook",
            filename="statement.xls",
            mime_type="application/vnd.ms-excel",
            bank_account_id="bank-1",
        )


def test_extract_statement_routes_xlsx_without_changing_response_shape():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Description", "Amount"])
    sheet.append([date(2026, 4, 1), "Receipt", 75])

    header, lines = facade.extract_statement(
        _workbook_bytes(workbook),
        filename="statement.xlsx",
        mime_type=XLSX_MIME,
        bank_account_id="bank-1",
        currency="ZAR",
    )

    assert header["source_format"] == "xlsx"
    assert header["parser_strategy"] == "deterministic_xlsx"
    assert header["extractor_type"] == "bank_statement"
    assert len(lines) == 1


def test_compatibility_facade_reexports_extraction_public_api():
    assert facade.ParsedBankLine is extraction.ParsedBankLine
    assert facade.extract_statement is extraction.extract_statement
    assert facade.parse_csv_statement is extraction.parse_csv_statement
    assert facade.parse_xlsx_statement is extraction.parse_xlsx_statement
    assert facade.money is extraction.money
