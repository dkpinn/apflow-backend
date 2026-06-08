from __future__ import annotations

import calendar
import csv
import io
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Optional


ZERO = Decimal("0.00")
MONEY = Decimal("0.01")
DEBIT_NORMAL_TYPES = {"asset", "expense", "other"}
CREDIT_NORMAL_TYPES = {"income", "liability", "equity"}
CURRENT_FY_TYPES = {"income", "expense"}

MONTH_NAME_TO_NUM = {name: index for index, name in enumerate(calendar.month_name) if name}
DEFAULT_FINANCIAL_YEAR_END = "February"


def money(value: Any) -> Decimal:
    if value in (None, ""):
        return ZERO
    try:
        return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)
    except Exception:
        return ZERO


def amount_out(value: Decimal) -> float:
    return float(value.quantize(MONEY, rounding=ROUND_HALF_UP))


def _fetch_rows(query) -> list[dict]:
    result = query.execute()
    return list(result.data or [])


def _parse_date(value: str, *, field: str = "as_at_date") -> date:
    try:
        return date.fromisoformat(str(value))
    except (ValueError, TypeError) as exc:
        raise ValueError(f"{field} must use YYYY-MM-DD format") from exc


def _last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def _financial_year_bounds(year_end_month: str, reference_date: date) -> tuple[date, date]:
    end_month = MONTH_NAME_TO_NUM.get((year_end_month or DEFAULT_FINANCIAL_YEAR_END).strip().title())
    if not end_month:
        end_month = MONTH_NAME_TO_NUM[DEFAULT_FINANCIAL_YEAR_END]

    if end_month == 12:
        return date(reference_date.year, 1, 1), date(reference_date.year, 12, 31)

    start_month = end_month + 1
    if reference_date.month >= start_month:
        start_year, end_year = reference_date.year, reference_date.year + 1
    else:
        start_year, end_year = reference_date.year - 1, reference_date.year

    fy_start = date(start_year, start_month, 1)
    fy_end = date(end_year, end_month, _last_day_of_month(end_year, end_month))
    return fy_start, fy_end


def _equivalent_date(reference_date: date, year: int) -> date:
    day = min(reference_date.day, _last_day_of_month(year, reference_date.month))
    return date(year, reference_date.month, day)


def _fetch_organisation_financial_year_end(db, organisation_id: str) -> Optional[str]:
    rows = _fetch_rows(
        db.table("organisations")
        .select("financial_year_end")
        .eq("id", organisation_id)
        .limit(1)
    )
    return rows[0].get("financial_year_end") if rows else None


def resolve_financial_year_end(db, *, organisation_id: str, override: Optional[str] = None) -> str:
    if override and override.strip():
        candidate = override.strip().title()
        if candidate not in MONTH_NAME_TO_NUM:
            raise ValueError("financial_year_end must be a full month name")
        return candidate
    saved = _fetch_organisation_financial_year_end(db, organisation_id)
    if saved and saved.strip().title() in MONTH_NAME_TO_NUM:
        return saved.strip().title()
    return DEFAULT_FINANCIAL_YEAR_END


def _fetch_accounts(db, organisation_id: str) -> list[dict]:
    return _fetch_rows(
        db.table("accounts")
        .select("id, code, name, type, group_name, active")
        .eq("organisation_id", organisation_id)
        .order("type")
        .order("code", desc=False, nullsfirst=True)
    )


def _fetch_posted_journals_until(db, organisation_id: str, cutoff: date) -> list[dict]:
    return _fetch_rows(
        db.table("gl_journals")
        .select("id, journal_date")
        .eq("organisation_id", organisation_id)
        .eq("status", "posted")
        .lte("journal_date", cutoff.isoformat())
    )


def _fetch_journal_lines(db, organisation_id: str, journal_ids: list[str]) -> list[dict]:
    if not journal_ids:
        return []
    return _fetch_rows(
        db.table("gl_journal_lines")
        .select("id, gl_journal_id, account_id, debit_amount, credit_amount, tracking")
        .eq("organisation_id", organisation_id)
        .in_("gl_journal_id", journal_ids)
    )


def _fetch_tracking_dimension(db, organisation_id: str, dimension_id: str) -> Optional[dict]:
    rows = _fetch_rows(
        db.table("tracking_dimensions")
        .select("id, name")
        .eq("organisation_id", organisation_id)
        .eq("id", dimension_id)
        .limit(1)
    )
    return rows[0] if rows else None


def _fetch_tracking_values(db, dimension_id: str) -> list[dict]:
    return _fetch_rows(
        db.table("tracking_values")
        .select("id, code, name")
        .eq("dimension_id", dimension_id)
    )


def _fetch_account_budgets(db, organisation_id: str, account_ids: list[str]) -> list[dict]:
    if not account_ids:
        return []
    return _fetch_rows(
        db.table("account_budgets")
        .select("id, account_id, tracking_value_id, period_start, period_end, amount")
        .eq("organisation_id", organisation_id)
        .in_("account_id", account_ids)
    )


def _normal_balance_amount(account_type: str, debit_total: Decimal, credit_total: Decimal) -> Decimal:
    if account_type in CREDIT_NORMAL_TYPES:
        return credit_total - debit_total
    return debit_total - credit_total


def _prorated_budget(rows: list[dict], *, period_start: date, period_end: date) -> Decimal:
    total = ZERO
    for row in rows:
        try:
            row_start = _parse_date(row.get("period_start"), field="period_start")
            row_end = _parse_date(row.get("period_end"), field="period_end")
        except ValueError:
            continue
        overlap_start = max(row_start, period_start)
        overlap_end = min(row_end, period_end)
        if overlap_start > overlap_end:
            continue
        overlap_days = (overlap_end - overlap_start).days + 1
        total_days = (row_end - row_start).days + 1
        if total_days <= 0:
            continue
        share = (money(row.get("amount")) * overlap_days / total_days).quantize(MONEY, rounding=ROUND_HALF_UP)
        total += share
    return total


def _account_period_bounds(account_type: str, fy_start: date, as_at: date) -> Optional[date]:
    """Returns the lower date bound for summing journal lines for this account
    type — None for balance-sheet accounts (cumulative since inception), or the
    financial-year start for income/expense accounts (which reset each year)."""
    return fy_start if account_type in CURRENT_FY_TYPES else None


def _sum_lines(
    lines: list[dict],
    journal_dates: dict[str, date],
    *,
    date_from: Optional[date],
    date_to: date,
    tracking_dimension_id: Optional[str] = None,
) -> tuple[Decimal, Decimal, dict[str, tuple[Decimal, Decimal]]]:
    total_debit = ZERO
    total_credit = ZERO
    by_tracking: dict[str, tuple[Decimal, Decimal]] = {}
    for line in lines:
        line_date = journal_dates.get(str(line.get("gl_journal_id")))
        if line_date is None or line_date > date_to:
            continue
        if date_from is not None and line_date < date_from:
            continue
        debit = money(line.get("debit_amount"))
        credit = money(line.get("credit_amount"))
        total_debit += debit
        total_credit += credit
        if tracking_dimension_id:
            tracking = line.get("tracking") if isinstance(line.get("tracking"), dict) else {}
            key = str(tracking.get(tracking_dimension_id) or "unassigned")
            existing_debit, existing_credit = by_tracking.get(key, (ZERO, ZERO))
            by_tracking[key] = (existing_debit + debit, existing_credit + credit)
    return total_debit, total_credit, by_tracking


def _net_balance(debit_total: Decimal, credit_total: Decimal) -> tuple[Decimal, Decimal]:
    net = debit_total - credit_total
    if net >= ZERO:
        return net, ZERO
    return ZERO, -net


def generate_trial_balance(
    db,
    *,
    organisation_id: str,
    as_at_date: str,
    financial_year_end: Optional[str] = None,
    tracking_dimension_id: Optional[str] = None,
    compare_years: Optional[list[int]] = None,
    include_budget: bool = False,
) -> dict:
    as_at = _parse_date(as_at_date)
    resolved_year_end = resolve_financial_year_end(db, organisation_id=organisation_id, override=financial_year_end)
    fy_start, fy_end = _financial_year_bounds(resolved_year_end, as_at)

    warnings: list[dict] = []
    compare_years = sorted(set(compare_years or []))

    periods: list[dict] = [{"key": "current", "year": as_at.year, "as_at": as_at, "fy_start": fy_start, "fy_end": fy_end}]
    for year in compare_years:
        comp_as_at = _equivalent_date(as_at, year)
        comp_fy_start, comp_fy_end = _financial_year_bounds(resolved_year_end, comp_as_at)
        periods.append({
            "key": f"compare_{year}",
            "year": year,
            "as_at": comp_as_at,
            "fy_start": comp_fy_start,
            "fy_end": comp_fy_end,
        })

    accounts = _fetch_accounts(db, organisation_id)
    accounts_by_id = {str(row.get("id")): row for row in accounts if row.get("id")}
    account_ids = list(accounts_by_id.keys())

    cutoff = max(period["as_at"] for period in periods)
    journals = _fetch_posted_journals_until(db, organisation_id, cutoff)
    journal_dates: dict[str, date] = {}
    for row in journals:
        journal_id = row.get("id")
        if not journal_id or not row.get("journal_date"):
            continue
        try:
            journal_dates[str(journal_id)] = _parse_date(row["journal_date"], field="journal_date")
        except ValueError:
            continue
    journal_lines = _fetch_journal_lines(db, organisation_id, list(journal_dates.keys()))

    lines_by_account: dict[str, list[dict]] = {}
    unmapped_lines = 0
    for line in journal_lines:
        account_id = line.get("account_id")
        if not account_id or str(account_id) not in accounts_by_id:
            unmapped_lines += 1
            continue
        lines_by_account.setdefault(str(account_id), []).append(line)
    if unmapped_lines:
        warnings.append({
            "code": "missing_account",
            "message": f"{unmapped_lines} journal line(s) reference an account that could not be loaded and were excluded.",
        })

    tracking_dimension = None
    tracking_value_labels: dict[str, str] = {}
    if tracking_dimension_id:
        tracking_dimension = _fetch_tracking_dimension(db, organisation_id, tracking_dimension_id)
        if not tracking_dimension:
            raise ValueError("Tracking dimension not found for this organisation")
        for value in _fetch_tracking_values(db, tracking_dimension_id):
            value_id = value.get("id")
            if value_id:
                tracking_value_labels[str(value_id)] = value.get("name") or value.get("code") or str(value_id)
        tracking_value_labels["unassigned"] = "Unassigned"

    budgets_by_account: dict[str, list[dict]] = {}
    if include_budget:
        for row in _fetch_account_budgets(db, organisation_id, account_ids):
            account_id = row.get("account_id")
            if account_id:
                budgets_by_account.setdefault(str(account_id), []).append(row)

    period_results: dict[str, dict] = {}
    for period in periods:
        accounts_out: list[dict] = []
        period_total_debit = ZERO
        period_total_credit = ZERO
        for account_id in account_ids:
            account = accounts_by_id[account_id]
            account_type = account.get("type") or "other"
            lines = lines_by_account.get(account_id, [])
            date_from = _account_period_bounds(account_type, period["fy_start"], period["as_at"])
            debit_total, credit_total, by_tracking = _sum_lines(
                lines,
                journal_dates,
                date_from=date_from,
                date_to=period["as_at"],
                tracking_dimension_id=tracking_dimension_id,
            )
            if not debit_total and not credit_total and not lines:
                continue

            debit_balance, credit_balance = _net_balance(debit_total, credit_total)
            period_total_debit += debit_balance
            period_total_credit += credit_balance

            entry: dict[str, Any] = {
                "account_id": account_id,
                "code": account.get("code"),
                "name": account.get("name"),
                "type": account_type,
                "group_name": account.get("group_name"),
                "debit": amount_out(debit_balance),
                "credit": amount_out(credit_balance),
            }

            if tracking_dimension_id:
                breakdown = []
                for value_id, (value_debit, value_credit) in by_tracking.items():
                    value_debit_balance, value_credit_balance = _net_balance(value_debit, value_credit)
                    breakdown.append({
                        "tracking_value_id": None if value_id == "unassigned" else value_id,
                        "label": tracking_value_labels.get(value_id, value_id),
                        "debit": amount_out(value_debit_balance),
                        "credit": amount_out(value_credit_balance),
                    })
                breakdown.sort(key=lambda item: item["label"].lower())
                entry["tracking_breakdown"] = breakdown

            if period["key"] == "current" and include_budget:
                budget_rows = budgets_by_account.get(account_id, [])
                budget_amount = _prorated_budget(budget_rows, period_start=period["fy_start"], period_end=period["as_at"])
                actual_amount = _normal_balance_amount(account_type, debit_total, credit_total)
                entry["budget"] = amount_out(budget_amount)
                entry["budget_actual"] = amount_out(actual_amount)
                entry["budget_variance"] = amount_out(actual_amount - budget_amount)

            accounts_out.append(entry)

        period_results[period["key"]] = {
            "year": period["year"],
            "as_at_date": period["as_at"].isoformat(),
            "financial_year": {
                "start": period["fy_start"].isoformat(),
                "end": period["fy_end"].isoformat(),
            },
            "accounts": accounts_out,
            "summary": {
                "total_debit": amount_out(period_total_debit),
                "total_credit": amount_out(period_total_credit),
                "in_balance": period_total_debit == period_total_credit,
            },
        }

    current = period_results.pop("current")
    comparisons = [period_results[f"compare_{year}"] for year in compare_years]

    if not current["summary"]["in_balance"]:
        warnings.append({
            "code": "out_of_balance",
            "message": "Total debits do not equal total credits for the selected date.",
            "total_debit": current["summary"]["total_debit"],
            "total_credit": current["summary"]["total_credit"],
        })

    return {
        "organisation_id": organisation_id,
        "as_at_date": current["as_at_date"],
        "financial_year_end": resolved_year_end,
        "financial_year": current["financial_year"],
        "tracking_dimension": (
            {"id": str(tracking_dimension.get("id")), "name": tracking_dimension.get("name")}
            if tracking_dimension else None
        ),
        "compare_years": compare_years,
        "budget_included": bool(include_budget),
        "accounts": current["accounts"],
        "summary": current["summary"],
        "comparisons": comparisons,
        "warnings": warnings,
        "disclaimer": (
            "Calculated from posted general-ledger journal entries as at the selected date. "
            "Balance-sheet accounts show cumulative balances; income and expense accounts show "
            "year-to-date movements since the start of the financial year."
        ),
    }


EXPORT_COLUMNS = [
    ("code", "Account Code"),
    ("name", "Account Name"),
    ("type", "Type"),
    ("debit", "Debit"),
    ("credit", "Credit"),
]


def trial_balance_csv(report: dict) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for row in report.get("accounts") or []:
        writer.writerow([row.get(key, "") for key, _ in EXPORT_COLUMNS])
    writer.writerow([])
    summary = report.get("summary") or {}
    writer.writerow(["", "", "Total", summary.get("total_debit", 0), summary.get("total_credit", 0)])
    return stream.getvalue().encode("utf-8-sig")


def trial_balance_text(report: dict) -> bytes:
    stream = io.StringIO()
    summary = report.get("summary") or {}
    stream.write("Trial Balance\n")
    stream.write(f"As at\t{report.get('as_at_date')}\n")
    stream.write(f"Financial year\t{(report.get('financial_year') or {}).get('start')}\tto\t{(report.get('financial_year') or {}).get('end')}\n")
    stream.write(f"Total debit\t{summary.get('total_debit', 0):.2f}\n")
    stream.write(f"Total credit\t{summary.get('total_credit', 0):.2f}\n")
    stream.write(f"In balance\t{summary.get('in_balance', False)}\n")
    stream.write(f"Notice\t{report.get('disclaimer', '')}\n\n")
    writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for row in report.get("accounts") or []:
        writer.writerow([row.get(key, "") for key, _ in EXPORT_COLUMNS])
    return stream.getvalue().encode("utf-8")


def trial_balance_xlsx(report: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel exports") from exc

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Trial Balance", ""])
    summary_sheet.append(["As at", date.fromisoformat(str(report.get("as_at_date")))])
    financial_year = report.get("financial_year") or {}
    summary_sheet.append(["Financial year start", date.fromisoformat(str(financial_year.get("start")))])
    summary_sheet.append(["Financial year end", date.fromisoformat(str(financial_year.get("end")))])
    summary_sheet.append(["Notice", report.get("disclaimer")])
    summary_sheet.append([])
    summary = report.get("summary") or {}
    summary_sheet.append(["Total debit", summary.get("total_debit", 0)])
    summary_sheet.append(["Total credit", summary.get("total_credit", 0)])
    summary_sheet.append(["In balance", "Yes" if summary.get("in_balance") else "No"])
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet.column_dimensions["A"].width = 26
    summary_sheet.column_dimensions["B"].width = 22
    for cell in ("B2", "B3", "B4"):
        summary_sheet[cell].number_format = "yyyy-mm-dd"
    for cell in summary_sheet["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    detail_sheet = workbook.create_sheet("Trial Balance")
    detail_sheet.append([label for _, label in EXPORT_COLUMNS])
    for row in report.get("accounts") or []:
        detail_sheet.append([row.get(key, "") for key, _ in EXPORT_COLUMNS])
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    widths = [16, 36, 14, 16, 16]
    for index, width in enumerate(widths, start=1):
        detail_sheet.column_dimensions[chr(64 + index)].width = width
    for row in detail_sheet.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
