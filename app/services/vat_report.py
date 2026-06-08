from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from typing import Any


ZERO = Decimal("0.00")
MONEY = Decimal("0.01")
VAT_TREATMENTS = {"full", "blocked", "exempt", "zero_rated"}


def money(value: Any) -> Decimal:
    if value in (None, ""):
        return ZERO
    try:
        return Decimal(str(value)).quantize(MONEY, rounding=ROUND_HALF_UP)
    except Exception:
        return ZERO


def amount_out(value: Decimal) -> float:
    return float(value.quantize(MONEY, rounding=ROUND_HALF_UP))


def allocate_amount_by_weights(total: Any, weights: list[Any]) -> list[Decimal]:
    amount = money(total)
    decimal_weights = [abs(money(weight)) for weight in weights]
    weight_total = sum(decimal_weights, ZERO)
    if not decimal_weights:
        return []
    if not weight_total:
        return [ZERO for _ in decimal_weights]

    shares: list[Decimal] = []
    allocated = ZERO
    for index, weight in enumerate(decimal_weights):
        if index == len(decimal_weights) - 1:
            share = amount - allocated
        else:
            share = (amount * weight / weight_total).quantize(MONEY, rounding=ROUND_HALF_UP)
        shares.append(share)
        allocated += share
    return shares


def _fetch_rows(query) -> list[dict]:
    result = query.execute()
    return list(result.data or [])


def _validate_dates(date_from: str, date_to: str) -> None:
    try:
        start = date.fromisoformat(date_from)
        end = date.fromisoformat(date_to)
    except ValueError as exc:
        raise ValueError("Dates must use YYYY-MM-DD format") from exc
    if start > end:
        raise ValueError("From date must be on or before To date")


def allocate_invoice_vat(
    *,
    invoice_tax: Any,
    line_items: list[dict],
    supplier_has_vat_number: bool,
) -> dict:
    """
    Allocate invoice VAT to saved lines for posting and historical reporting.

    Explicit line tax is respected first. Any remaining invoice VAT is allocated
    proportionally across standard-rated lines that do not have explicit tax.
    Non-claimable residual VAT is expensed so the journal remains balanced.
    """
    total_tax = money(invoice_tax)
    allocations: list[dict] = []

    for index, line in enumerate(line_items or []):
        treatment = str(line.get("vat_treatment") or "full").lower()
        if treatment not in VAT_TREATMENTS:
            treatment = "full"
        explicit_raw = line.get("tax_amount")
        explicit = money(explicit_raw) if explicit_raw not in (None, "") else None
        allocations.append({
            "index": index,
            "line_id": line.get("id"),
            "treatment": treatment,
            "base_amount": abs(money(line.get("line_total"))),
            "explicit_tax": explicit,
            "allocated_tax": ZERO,
            "claimable_tax": ZERO,
            "blocked_tax": ZERO,
        })

    eligible = [row for row in allocations if row["treatment"] in {"full", "blocked"}]
    explicit_total = sum(
        (row["explicit_tax"] or ZERO for row in eligible),
        ZERO,
    )
    remaining = max(total_tax - explicit_total, ZERO)

    explicit_rows = [row for row in eligible if row["explicit_tax"] is not None]
    if explicit_total > total_tax and explicit_total:
        scaled = allocate_amount_by_weights(
            total_tax,
            [row["explicit_tax"] for row in explicit_rows],
        )
        for row, share in zip(explicit_rows, scaled):
            row["allocated_tax"] = share
        remaining = ZERO
    else:
        for row in explicit_rows:
            row["allocated_tax"] = row["explicit_tax"]

    missing = [row for row in eligible if row["explicit_tax"] is None]
    missing_weight = sum((row["base_amount"] for row in missing), ZERO)
    allocated_remaining = ZERO
    for position, row in enumerate(missing):
        if position == len(missing) - 1:
            share = remaining - allocated_remaining
        elif missing_weight:
            share = (remaining * row["base_amount"] / missing_weight).quantize(MONEY, rounding=ROUND_HALF_UP)
        else:
            share = ZERO
        row["allocated_tax"] = share
        allocated_remaining += share

    allocated_total = sum((row["allocated_tax"] for row in allocations), ZERO)
    unallocated_tax = max(total_tax - allocated_total, ZERO)

    if unallocated_tax:
        expense_targets = [row for row in allocations if row["base_amount"] > ZERO] or allocations
        target_weight = sum((row["base_amount"] for row in expense_targets), ZERO)
        distributed = ZERO
        for position, row in enumerate(expense_targets):
            if position == len(expense_targets) - 1:
                share = unallocated_tax - distributed
            elif target_weight:
                share = (unallocated_tax * row["base_amount"] / target_weight).quantize(MONEY, rounding=ROUND_HALF_UP)
            else:
                share = ZERO
            row["allocated_tax"] += share
            distributed += share

    for row in allocations:
        if supplier_has_vat_number and row["treatment"] == "full":
            row["claimable_tax"] = row["allocated_tax"]
        else:
            row["blocked_tax"] = row["allocated_tax"]

    claimable = sum((row["claimable_tax"] for row in allocations), ZERO)
    blocked = sum((row["blocked_tax"] for row in allocations), ZERO)
    return {
        "line_allocations": allocations,
        "claimable_tax": claimable,
        "blocked_tax": blocked,
        "invoice_tax": total_tax,
    }


def _invoice_claimability(invoice: dict, line_items: list[dict], supplier: dict | None) -> dict:
    supplier_vat = (
        (supplier or {}).get("vat_number")
        or invoice.get("vat_number_extracted")
        or ""
    )
    return allocate_invoice_vat(
        invoice_tax=invoice.get("tax_amount"),
        line_items=line_items,
        supplier_has_vat_number=bool(str(supplier_vat).strip()),
    )


def _invoice_metadata(invoice: dict, supplier: dict | None) -> dict:
    supplier_name = (
        (supplier or {}).get("supplier_name")
        or (supplier or {}).get("trading_name")
        or invoice.get("supplier_name_extracted")
        or ""
    )
    supplier_vat = (
        (supplier or {}).get("vat_number")
        or invoice.get("vat_number_extracted")
        or ""
    )
    gross = invoice.get("total_amount")
    if gross in (None, ""):
        gross = money(invoice.get("subtotal")) + money(invoice.get("tax_amount"))
    return {
        "supplier": supplier_name,
        "supplier_vat_number": supplier_vat,
        "invoice_number": invoice.get("invoice_number") or "",
        "gross_amount": amount_out(money(gross)),
    }


def generate_vat_report(
    db,
    *,
    organisation_id: str,
    date_from: str,
    date_to: str,
) -> dict:
    _validate_dates(date_from, date_to)

    vat_accounts = _fetch_rows(
        db.table("accounts")
        .select("id, code, name")
        .eq("organisation_id", organisation_id)
        .eq("system_key", "vat_control")
        .limit(1)
    )
    if not vat_accounts:
        raise ValueError("VAT Control account is not configured for this organisation")
    vat_account = vat_accounts[0]

    journals = _fetch_rows(
        db.table("gl_journals")
        .select("id, source_type, source_id, journal_date, description, created_at")
        .eq("organisation_id", organisation_id)
        .eq("status", "posted")
        .lte("journal_date", date_to)
        .order("journal_date")
        .order("created_at")
    )
    journal_by_id = {str(row.get("id")): row for row in journals if row.get("id")}
    journal_ids = list(journal_by_id)

    vat_lines = []
    if journal_ids:
        vat_lines = _fetch_rows(
            db.table("gl_journal_lines")
            .select("id, gl_journal_id, description, debit_amount, credit_amount, created_at, sort_order")
            .eq("organisation_id", organisation_id)
            .eq("account_id", vat_account.get("id"))
            .in_("gl_journal_id", journal_ids)
            .order("sort_order")
            .order("created_at")
        )

    invoice_ids = sorted({
        str(journal.get("source_id"))
        for journal in journals
        if journal.get("source_type") == "invoice" and journal.get("source_id")
    })
    invoices: list[dict] = []
    if invoice_ids:
        invoices = _fetch_rows(
            db.table("invoices_extracted")
            .select(
                "id, supplier_id, supplier_name_extracted, vat_number_extracted, "
                "invoice_number, subtotal, tax_amount, total_amount"
            )
            .eq("organisation_id", organisation_id)
            .in_("id", invoice_ids)
        )
    invoice_by_id = {str(row.get("id")): row for row in invoices if row.get("id")}

    supplier_ids = sorted({str(row.get("supplier_id")) for row in invoices if row.get("supplier_id")})
    suppliers: list[dict] = []
    if supplier_ids:
        suppliers = _fetch_rows(
            db.table("suppliers")
            .select("id, supplier_name, trading_name, vat_number")
            .eq("organisation_id", organisation_id)
            .in_("id", supplier_ids)
        )
    supplier_by_id = {str(row.get("id")): row for row in suppliers if row.get("id")}

    line_items: list[dict] = []
    if invoice_ids:
        line_items = _fetch_rows(
            db.table("invoice_line_items")
            .select("id, invoice_extracted_id, line_total, tax_amount, vat_treatment")
            .eq("organisation_id", organisation_id)
            .in_("invoice_extracted_id", invoice_ids)
        )
    items_by_invoice: dict[str, list[dict]] = {}
    for item in line_items:
        items_by_invoice.setdefault(str(item.get("invoice_extracted_id")), []).append(item)

    claimability_by_invoice: dict[str, dict] = {}
    metadata_by_invoice: dict[str, dict] = {}
    for invoice_id, invoice in invoice_by_id.items():
        supplier = supplier_by_id.get(str(invoice.get("supplier_id")))
        claimability_by_invoice[invoice_id] = _invoice_claimability(
            invoice,
            items_by_invoice.get(invoice_id, []),
            supplier,
        )
        metadata_by_invoice[invoice_id] = _invoice_metadata(invoice, supplier)

    period_invoice_ids = {
        str(journal.get("source_id"))
        for journal in journals
        if journal.get("source_type") == "invoice"
        and journal.get("source_id")
        and date_from <= str(journal.get("journal_date") or "") <= date_to
    }

    opening_balance = ZERO
    output_vat = ZERO
    posted_input_vat = ZERO
    allowable_input_vat = sum(
        (
            money((claimability_by_invoice.get(invoice_id) or {}).get("claimable_tax"))
            for invoice_id in period_invoice_ids
        ),
        ZERO,
    )
    blocked_input_vat = sum(
        (
            money((claimability_by_invoice.get(invoice_id) or {}).get("blocked_tax"))
            for invoice_id in period_invoice_ids
        ),
        ZERO,
    )
    historical_variance = ZERO
    warnings: list[dict] = []
    period_rows: list[dict] = []
    invoice_vat_debits: dict[str, Decimal] = {}
    detailed_invoice_claimability: set[str] = set()

    sorted_lines = sorted(
        vat_lines,
        key=lambda row: (
            str((journal_by_id.get(str(row.get("gl_journal_id"))) or {}).get("journal_date") or ""),
            str((journal_by_id.get(str(row.get("gl_journal_id"))) or {}).get("created_at") or ""),
            int(row.get("sort_order") or 0),
            str(row.get("created_at") or ""),
            str(row.get("id") or ""),
        ),
    )

    for line in sorted_lines:
        journal = journal_by_id.get(str(line.get("gl_journal_id"))) or {}
        journal_date = str(journal.get("journal_date") or "")
        debit = money(line.get("debit_amount"))
        credit = money(line.get("credit_amount"))
        movement = credit - debit
        if journal_date < date_from:
            opening_balance += movement
            continue

        source_invoice_id = (
            str(journal.get("source_id"))
            if journal.get("source_type") == "invoice" and journal.get("source_id")
            else None
        )
        metadata = metadata_by_invoice.get(source_invoice_id or "", {})
        claimability = claimability_by_invoice.get(source_invoice_id or "")

        row_allowable = debit
        row_blocked = ZERO
        row_variance = ZERO
        if claimability:
            invoice_vat_debits[source_invoice_id] = (
                invoice_vat_debits.get(source_invoice_id, ZERO) + debit
            )
            if source_invoice_id not in detailed_invoice_claimability:
                row_allowable = claimability["claimable_tax"]
                row_blocked = claimability["blocked_tax"]
                detailed_invoice_claimability.add(source_invoice_id)
            else:
                row_allowable = ZERO
            row_variance = debit - row_allowable
        else:
            allowable_input_vat += debit

        output_vat += credit
        posted_input_vat += debit

        period_rows.append({
            "id": line.get("id"),
            "journal_id": line.get("gl_journal_id"),
            "date": journal_date,
            "supplier": metadata.get("supplier") or "",
            "supplier_vat_number": metadata.get("supplier_vat_number") or "",
            "invoice_number": metadata.get("invoice_number") or "",
            "description": line.get("description") or journal.get("description") or "",
            "debit_amount": amount_out(debit),
            "credit_amount": amount_out(credit),
            "gross_amount": metadata.get("gross_amount") or amount_out(max(debit, credit)),
            "allowable_input_vat": amount_out(row_allowable),
            "blocked_input_vat": amount_out(row_blocked),
            "claimability_variance": amount_out(row_variance),
        })

    for invoice_id in period_invoice_ids:
        actual_debit = invoice_vat_debits.get(invoice_id, ZERO)
        claimable = money(
            (claimability_by_invoice.get(invoice_id) or {}).get("claimable_tax")
        )
        historical_variance += max(actual_debit - claimable, ZERO)

    running = opening_balance
    for row in period_rows:
        running += money(row["credit_amount"]) - money(row["debit_amount"])
        row["running_total"] = amount_out(running)

    period_movement = output_vat - posted_input_vat
    actual_closing_balance = opening_balance + period_movement
    calculated_vat_position = opening_balance + output_vat - allowable_input_vat

    if historical_variance > ZERO:
        warnings.append({
            "code": "historical_claimability_variance",
            "message": (
                "Historical invoice postings include VAT that is not currently claimable. "
                "The adjusted calculated position differs from the VAT Control balance."
            ),
            "amount": amount_out(historical_variance),
        })
    if not any(row.get("credit_amount") for row in period_rows):
        warnings.append({
            "code": "no_output_vat_detail",
            "message": (
                "No output VAT credits were posted in this period. Output VAT is limited "
                "to transactions already posted to the VAT Control account."
            ),
        })

    return {
        "organisation_id": organisation_id,
        "date_from": date_from,
        "date_to": date_to,
        "vat_control_account": vat_account,
        "summary": {
            "opening_balance": amount_out(opening_balance),
            "output_vat": amount_out(output_vat),
            "posted_input_vat": amount_out(posted_input_vat),
            "allowable_input_vat": amount_out(allowable_input_vat),
            "blocked_input_vat": amount_out(blocked_input_vat),
            "period_vat_payable_refundable": amount_out(output_vat - allowable_input_vat),
            "period_gl_movement": amount_out(period_movement),
            "closing_vat_control_balance": amount_out(actual_closing_balance),
            "calculated_vat_position": amount_out(calculated_vat_position),
            "historical_claimability_variance": amount_out(historical_variance),
        },
        "rows": period_rows,
        "warnings": warnings,
        "disclaimer": (
            "Calculated from posted accounting entries. This is not a SARS eFiling "
            "statement or confirmed SARS account balance."
        ),
    }


EXPORT_COLUMNS = [
    ("date", "Date"),
    ("supplier", "Supplier"),
    ("supplier_vat_number", "Supplier VAT Number"),
    ("invoice_number", "Invoice Number"),
    ("description", "Description / Details"),
    ("debit_amount", "Debit Amount"),
    ("credit_amount", "Credit Amount"),
    ("gross_amount", "Gross Amount"),
    ("running_total", "Running Total"),
]


def vat_report_csv(report: dict) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for row in report.get("rows") or []:
        writer.writerow([row.get(key, "") for key, _ in EXPORT_COLUMNS])
    return stream.getvalue().encode("utf-8-sig")


def vat_report_text(report: dict) -> bytes:
    stream = io.StringIO()
    summary = report.get("summary") or {}
    stream.write("VAT Report\n")
    stream.write(f"Period\t{report.get('date_from')}\t{report.get('date_to')}\n")
    stream.write(f"Calculated VAT position\t{summary.get('calculated_vat_position', 0):.2f}\n")
    stream.write(f"Output VAT\t{summary.get('output_vat', 0):.2f}\n")
    stream.write(f"Allowable input VAT\t{summary.get('allowable_input_vat', 0):.2f}\n")
    stream.write(f"Blocked input VAT\t{summary.get('blocked_input_vat', 0):.2f}\n")
    stream.write(f"Closing VAT Control balance\t{summary.get('closing_vat_control_balance', 0):.2f}\n")
    stream.write(f"Notice\t{report.get('disclaimer', '')}\n\n")
    writer = csv.writer(stream, delimiter="\t", lineterminator="\n")
    writer.writerow([label for _, label in EXPORT_COLUMNS])
    for row in report.get("rows") or []:
        writer.writerow([row.get(key, "") for key, _ in EXPORT_COLUMNS])
    return stream.getvalue().encode("utf-8")


def vat_report_xlsx(report: dict) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel exports") from exc

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["VAT Report", ""])
    summary_sheet.append(["From", date.fromisoformat(str(report.get("date_from")))])
    summary_sheet.append(["To", date.fromisoformat(str(report.get("date_to")))])
    summary_sheet.append(["Notice", report.get("disclaimer")])
    summary_sheet.append([])
    labels = {
        "opening_balance": "Opening VAT Control balance",
        "output_vat": "Output VAT",
        "posted_input_vat": "Posted input VAT",
        "allowable_input_vat": "Allowable input VAT",
        "blocked_input_vat": "Blocked input VAT",
        "period_vat_payable_refundable": "Period VAT payable / refundable",
        "closing_vat_control_balance": "Closing VAT Control balance",
        "calculated_vat_position": "Calculated VAT position",
        "historical_claimability_variance": "Historical claimability variance",
    }
    for key, label in labels.items():
        summary_sheet.append([label, (report.get("summary") or {}).get(key, 0)])
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet.column_dimensions["A"].width = 34
    summary_sheet.column_dimensions["B"].width = 22
    summary_sheet["B2"].number_format = "yyyy-mm-dd"
    summary_sheet["B3"].number_format = "yyyy-mm-dd"
    for cell in summary_sheet["B"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

    detail_sheet = workbook.create_sheet("VAT Detail")
    detail_sheet.append([label for _, label in EXPORT_COLUMNS])
    for row in report.get("rows") or []:
        values = [row.get(key, "") for key, _ in EXPORT_COLUMNS]
        if values[0]:
            values[0] = date.fromisoformat(str(values[0]))
        detail_sheet.append(values)
    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    widths = [13, 25, 20, 18, 42, 16, 16, 16, 16]
    for index, width in enumerate(widths, start=1):
        detail_sheet.column_dimensions[chr(64 + index)].width = width
    for row in detail_sheet.iter_rows(min_row=2, min_col=6, max_col=9):
        for cell in row:
            cell.number_format = '#,##0.00;[Red]-#,##0.00'
    for cell in detail_sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
