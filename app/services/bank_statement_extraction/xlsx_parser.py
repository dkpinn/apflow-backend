from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import load_workbook

from .models import ParsedBankLine
from .tabular import parse_tabular_rows, recognizable_header_score


MAX_HEADER_SEARCH_ROWS = 50


def _header_values(row: tuple[Any, ...]) -> list[str]:
    return [str(value).strip() if value is not None else "" for value in row]


def parse_xlsx_statement(
    file_bytes: bytes,
    *,
    bank_account_id: str,
    currency: Optional[str] = None,
) -> tuple[dict[str, Any], list[ParsedBankLine]]:
    try:
        workbook = load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError("Could not read XLSX bank statement; the workbook may be corrupt or unsupported") from exc

    try:
        best: tuple[int, int, int, Any, list[str]] | None = None
        for sheet_index, worksheet in enumerate(workbook.worksheets):
            if worksheet.sheet_state != "visible":
                continue
            for header_row_index, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SEARCH_ROWS, values_only=True),
                start=1,
            ):
                fieldnames = _header_values(row)
                score = recognizable_header_score(fieldnames)
                if not score:
                    continue
                candidate = (score, -sheet_index, -header_row_index, worksheet, fieldnames)
                if best is None or candidate[:3] > best[:3]:
                    best = candidate

        if best is None:
            raise ValueError("XLSX bank statement does not contain a recognizable transaction header")

        _score, _sheet_order, negative_header_row, worksheet, fieldnames = best
        header_row_index = -negative_header_row
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            if not any(value not in (None, "") for value in values):
                continue
            row = {
                fieldname: values[index] if index < len(values) else None
                for index, fieldname in enumerate(fieldnames)
                if fieldname
            }
            rows.append(row)

        return parse_tabular_rows(
            rows,
            fieldnames=[fieldname for fieldname in fieldnames if fieldname],
            bank_account_id=bank_account_id,
            currency=currency,
            source_format="xlsx",
            parser_strategy="deterministic_xlsx",
            confidence_score=0.98,
            metadata_extra={
                "sheet_name": worksheet.title,
                "header_row": header_row_index,
            },
        )
    finally:
        workbook.close()
